#!/usr/bin/env python3
"""
google_sheet_to_excel.py
Экспорт Google Sheets → Excel для журналов бетонных работ.

Использование:
    python google_sheet_to_excel.py --project 6.2

Требования:
    pip install gspread openpyxl
"""

import argparse
import os
import sys
from datetime import datetime

try:
    import gspread
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
except ImportError:
    print("Установите зависимости: pip install gspread openpyxl")
    sys.exit(1)


# === КОНФИГУРАЦИЯ ===
SPREADSHEET_IDS = {
    "6.2": {
        "concrete_journal": "14LATjCzXXCAIW6QpB6MMt12D-FeYHoScr6wDbMRHUVs",
        "care_journal": "13D8MDxsWPyBlk0JaSZk7ouMpZMvDkqCJuYtqdAyUS-4",
    },
    # Добавьте другие проекты по аналогии
}

OUTPUT_DIR = "/mnt/agents/output/exports"

# === СТИЛИ ===
HEADER_FONT = Font(bold=True, size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT_COLOR = Font(bold=True, size=11, color="FFFFFF")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def get_spreadsheet_data(sheet_id: str, sheet_name: str = "Лист1") -> list:
    """Получение данных из Google Sheets."""
    # Здесь должен быть ваш способ авторизации (service account или OAuth2)
    # gc = gspread.service_account(filename="credentials.json")
    # spreadsheet = gc.open_by_key(sheet_id)
    # worksheet = spreadsheet.worksheet(sheet_name)
    # return worksheet.get_all_records()

    # Заглушка для примера:
    print(f"[INFO] Получение данных из spreadsheet: {sheet_id}, sheet: {sheet_name}")
    return []


def create_excel(data: list, headers: list, output_path: str, title: str):
    """Создание Excel-файла с форматированием."""
    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]  # Excel ограничивает длину имени листа

    # Заголовки
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT_COLOR
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Данные
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            value = row_data.get(header, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical='center', wrap_text=True)

    # Автоширина колонок
    for col_idx, header in enumerate(headers, 1):
        max_length = len(str(header))
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[chr(64 + col_idx) if col_idx <= 26 else f"A{chr(64 + col_idx - 26)}"].width = adjusted_width

    # Закрепление заголовка
    ws.freeze_panes = "A2"

    wb.save(output_path)
    print(f"[OK] Файл сохранён: {output_path}")


def export_journal(project: str, journal_type: str):
    """Экспорт конкретного журнала."""
    if project not in SPREADSHEET_IDS:
        print(f"[ERROR] Проект '{project}' не найден в конфигурации")
        return

    sheet_id = SPREADSHEET_IDS[project][journal_type]
    data = get_spreadsheet_data(sheet_id)

    if not data:
        print(f"[WARN] Нет данных для {journal_type}")
        return

    headers = list(data[0].keys())
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{journal_type}_{project}_{timestamp}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, filename)

    os.makedirs(OUTPUT_DIR, exist_ok=True)
    create_excel(data, headers, output_path, f"{journal_type}_{project}")


def main():
    parser = argparse.ArgumentParser(description="Экспорт журналов бетонных работ в Excel")
    parser.add_argument("--project", required=True, help="Код проекта (например: 6.2)")
    parser.add_argument("--journal", choices=["concrete", "care", "all"], 
                        default="all", help="Тип журнала")
    args = parser.parse_args()

    print(f"=== Экспорт журналов для проекта {args.project} ===")

    if args.journal in ("concrete", "all"):
        export_journal(args.project, "concrete_journal")

    if args.journal in ("care", "all"):
        export_journal(args.project, "care_journal")

    print("=== Готово ===")


if __name__ == "__main__":
    main()
